from __future__ import annotations

import os
import sys
from collections import defaultdict
from datetime import datetime
from pathlib import Path

BASE_DIR = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(BASE_DIR))

os.environ.setdefault("DJANGO_SETTINGS_MODULE", "config.settings")

import django
from openpyxl import Workbook, load_workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

django.setup()

from main.models import (  # noqa: E402
    CatalogBlock,
    CatalogSystem,
    Direction,
    ProductGroup,
    ProductType,
    Vendor,
    VendorProductGroup,
)


OUT_DIR = BASE_DIR / "deliverables/catalog_vendor_review"

NAVY = "17243A"
EDIT = "FFF2CC"
WARN = "FCE4D6"
BORDER = "40516B"
WHITE = "FFFFFF"

THIN = Side(style="thin", color=BORDER)
HEADER_FILL = PatternFill("solid", fgColor=NAVY)
EDIT_FILL = PatternFill("solid", fgColor=EDIT)
WARN_FILL = PatternFill("solid", fgColor=WARN)
WHITE_FONT = Font(color=WHITE, bold=True)


def yesno(value: bool) -> str:
    return "Да" if value else "Нет"


def autosize(ws, max_width: int = 56) -> None:
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        width = 12
        for cell in col[:200]:
            if cell.value is None:
                continue
            lines = str(cell.value).split("\n")
            width = max(width, min(max(len(line) for line in lines) + 2, max_width))
        ws.column_dimensions[letter].width = width


def style_sheet(ws, freeze: str = "A2") -> None:
    ws.freeze_panes = freeze
    ws.sheet_view.showGridLines = False
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def write_table(ws, headers, rows, editable_cols=None, dropdowns=None) -> None:
    editable_cols = set(editable_cols or [])
    dropdowns = dropdowns or {}

    ws.append(headers)
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(1, col_idx)
        cell.fill = HEADER_FILL
        cell.font = WHITE_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for row in rows:
        ws.append(row)

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=ws.max_column):
        for cell in row:
            if cell.column in editable_cols:
                cell.fill = EDIT_FILL
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

    if ws.max_row >= 2:
        ws.auto_filter.ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"

    for col_idx in range(1, ws.max_column + 1):
        header = str(ws.cell(1, col_idx).value or "")
        if header in dropdowns and ws.max_row >= 2:
            validation = DataValidation(
                type="list",
                formula1=f'"{dropdowns[header]}"',
                allow_blank=True,
            )
            ws.add_data_validation(validation)
            validation.add(f"{get_column_letter(col_idx)}2:{get_column_letter(col_idx)}{ws.max_row}")

    ws.row_dimensions[1].height = 34
    style_sheet(ws)
    autosize(ws)


def main() -> None:
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    out_file = OUT_DIR / f"PNP_catalog_vendor_matrix_{stamp}.xlsx"

    status_values = "ОК,Добавить бренды,Убрать лишнее,Нужна проверка,На паузе"
    action_values = "Оставить,Добавить,Удалить,Проверить,Не показывать на сайте"
    priority_values = "P0,P1,P2,Пауза"

    wb = Workbook()
    ws_intro = wb.active
    ws_intro.title = "Инструкция"
    ws_matrix = wb.create_sheet("Матрица каталога")
    ws_links = wb.create_sheet("Связи группа-бренд")
    ws_vendors = wb.create_sheet("Производители")
    ws_empty = wb.create_sheet("Пустые группы")
    ws_summary = wb.create_sheet("Сводка")
    ws_review = wb.create_sheet("Проверка")

    blocks = list(CatalogBlock.objects.order_by("sort_order", "title"))
    product_groups = list(
        ProductGroup.objects.select_related("system__direction__block").order_by(
            "system__direction__block__sort_order",
            "system__direction__sort_order",
            "system__sort_order",
            "sort_order",
            "title",
        )
    )

    product_types_by_group = defaultdict(list)
    for product_type in ProductType.objects.select_related("product_group").order_by(
        "product_group_id", "sort_order", "title"
    ):
        product_types_by_group[product_type.product_group_id].append(product_type.title)

    links = list(
        VendorProductGroup.objects.select_related(
            "vendor", "product_group__system__direction__block"
        ).order_by(
            "product_group__system__direction__block__sort_order",
            "product_group__system__direction__sort_order",
            "product_group__system__sort_order",
            "product_group__sort_order",
            "vendor__name",
        )
    )
    links_by_group = defaultdict(list)
    links_by_vendor = defaultdict(list)
    for link in links:
        links_by_group[link.product_group_id].append(link)
        links_by_vendor[link.vendor_id].append(link)

    write_table(
        ws_intro,
        ["Раздел", "Описание"],
        [
            [
                "Назначение файла",
                "Навигация по каталогу ПНП: где какая товарная группа находится и какие производители сейчас привязаны.",
            ],
            [
                "Как работать",
                "1) Начните с листа «Пустые группы». 2) Потом проверьте «Матрица каталога». "
                "3) Все решения пишите в жёлтые колонки. 4) Бренды не придумываем: добавляем только подтверждённых производителей.",
            ],
            [
                "Главное правило",
                "Сайт должен оставаться актуальным: если бренд добавляется в каталог, он должен быть связан с конкретными товарными группами.",
            ],
            [
                "Колонки для коммерческого отдела",
                "Статус проверки, Приоритет, Производители к добавлению, Производители к удалению, Комментарий коммерческого отдела.",
            ],
            [
                "Технический статус",
                "Файл не меняет сайт и базу данных. После согласования разработчик переносит решения в базу.",
            ],
            ["Дата выгрузки", datetime.now().strftime("%d.%m.%Y %H:%M:%S")],
        ],
    )
    ws_intro.column_dimensions["A"].width = 30
    ws_intro.column_dimensions["B"].width = 110

    matrix_rows = []
    for group in product_groups:
        system = group.system
        direction = system.direction
        block = direction.block
        group_links = links_by_group.get(group.id, [])
        public_links = [
            link for link in group_links if link.show_in_catalog and link.status == "published"
        ]
        all_link_names = [
            f"{link.vendor.name} ({link.status}, catalog={yesno(link.show_in_catalog)})"
            for link in group_links
        ]
        status = "Нужна проверка" if not public_links else "ОК"
        priority = "P0" if not public_links else ""
        matrix_rows.append(
            [
                group.id,
                block.title,
                direction.title,
                system.title,
                group.title,
                f"{block.slug} / {direction.slug} / {system.slug} / {group.slug}",
                group.crm_category or "",
                "\n".join(product_types_by_group.get(group.id, [])),
                len(product_types_by_group.get(group.id, [])),
                "\n".join(link.vendor.name for link in public_links),
                len(public_links),
                "\n".join(all_link_names),
                status,
                priority,
                "",
                "",
                "",
                "",
            ]
        )

    write_table(
        ws_matrix,
        [
            "ID группы",
            "Глобальный блок",
            "Направление",
            "Система",
            "Товарная группа",
            "Навигационный путь",
            "CRM категория",
            "Типы продукции",
            "Кол-во типов",
            "Производители на сайте",
            "Кол-во производителей",
            "Все текущие связи",
            "Статус проверки",
            "Приоритет",
            "Производители к добавлению",
            "Производители к удалению",
            "Комментарий коммерческого отдела",
            "Комментарий для разработчика",
        ],
        matrix_rows,
        editable_cols={13, 14, 15, 16, 17, 18},
        dropdowns={"Статус проверки": status_values, "Приоритет": priority_values},
    )
    ws_matrix.conditional_formatting.add(
        f"K2:K{ws_matrix.max_row}", CellIsRule(operator="equal", formula=["0"], fill=WARN_FILL)
    )

    write_table(
        ws_links,
        [
            "ID связи",
            "ID группы",
            "Глобальный блок",
            "Направление",
            "Система",
            "Товарная группа",
            "Производитель",
            "Сайт производителя",
            "Статус связи",
            "Уверенность/источник",
            "Показывать в каталоге",
            "Показывать в производителях",
            "Показывать на главной",
            "Показывать в партнёрах",
            "Решение коммерческого отдела",
            "Комментарий коммерческого отдела",
            "Комментарий для разработчика",
        ],
        [
            [
                link.id,
                link.product_group.id,
                link.product_group.system.direction.block.title,
                link.product_group.system.direction.title,
                link.product_group.system.title,
                link.product_group.title,
                link.vendor.name,
                link.vendor.official_site or "",
                link.status,
                f"{link.confidence or ''}; {link.matched_by or ''}; {link.source or ''}",
                yesno(link.show_in_catalog),
                yesno(link.show_in_vendors),
                yesno(link.show_on_home),
                yesno(link.show_in_partners),
                "Оставить" if link.show_in_catalog and link.status == "published" else "Проверить",
                "",
                "",
            ]
            for link in links
        ],
        editable_cols={15, 16, 17},
        dropdowns={"Решение коммерческого отдела": action_values},
    )

    vendor_rows = []
    for vendor in Vendor.objects.order_by("name"):
        vendor_links = links_by_vendor.get(vendor.id, [])
        groups = sorted({link.product_group.title for link in vendor_links})
        vendor_rows.append(
            [
                vendor.id,
                vendor.name,
                vendor.slug,
                vendor.official_site or "",
                vendor.logo or "",
                vendor.status,
                vendor.confidence,
                "\n".join(sorted({link.product_group.system.direction.block.title for link in vendor_links})),
                "\n".join(sorted({link.product_group.system.direction.title for link in vendor_links})),
                "\n".join(sorted({link.product_group.system.title for link in vendor_links})),
                "\n".join(groups),
                len(groups),
                "Проверить" if vendor.status == "manual_review_needed" else "Оставить",
                "",
            ]
        )

    write_table(
        ws_vendors,
        [
            "ID производителя",
            "Производитель",
            "slug",
            "Официальный сайт",
            "Логотип",
            "Статус",
            "Уверенность",
            "Блоки",
            "Направления",
            "Системы",
            "Товарные группы",
            "Кол-во групп",
            "Решение коммерческого отдела",
            "Комментарий коммерческого отдела",
        ],
        vendor_rows,
        editable_cols={13, 14},
        dropdowns={"Решение коммерческого отдела": action_values},
    )

    empty_rows = [
        [
            row[0],
            row[1],
            row[2],
            row[3],
            row[4],
            row[7],
            row[6],
            "Подтвердить производителей или пометить, что группа временно без брендов",
            "P0",
            "",
            "",
        ]
        for row in matrix_rows
        if row[10] == 0
    ]
    write_table(
        ws_empty,
        [
            "ID группы",
            "Глобальный блок",
            "Направление",
            "Система",
            "Товарная группа",
            "Типы продукции",
            "CRM категория",
            "Что нужно от коммерческого отдела",
            "Приоритет",
            "Производители к добавлению",
            "Комментарий",
        ],
        empty_rows,
        editable_cols={8, 9, 10, 11},
        dropdowns={"Приоритет": priority_values},
    )

    summary_rows = []

    def add_summary(level, block="", direction="", system="", groups=None):
        groups = groups or []
        group_ids = {group.id for group in groups}
        types_count = sum(len(product_types_by_group.get(group.id, [])) for group in groups)
        vendors = {
            link.vendor.name
            for link in links
            if link.product_group_id in group_ids
            and link.show_in_catalog
            and link.status == "published"
        }
        empty_count = sum(
            1
            for group in groups
            if not [
                link
                for link in links_by_group.get(group.id, [])
                if link.show_in_catalog and link.status == "published"
            ]
        )
        summary_rows.append(
            [level, block, direction, system, len(groups), types_count, len(vendors), empty_count]
        )

    for block in blocks:
        block_groups = [group for group in product_groups if group.system.direction.block_id == block.id]
        add_summary("Блок", block.title, groups=block_groups)
        for direction in Direction.objects.filter(block=block).order_by("sort_order", "title"):
            direction_groups = [group for group in block_groups if group.system.direction_id == direction.id]
            add_summary("Направление", block.title, direction.title, groups=direction_groups)
            for system in CatalogSystem.objects.filter(direction=direction).order_by("sort_order", "title"):
                system_groups = [group for group in direction_groups if group.system_id == system.id]
                add_summary("Система", block.title, direction.title, system.title, system_groups)

    write_table(
        ws_summary,
        [
            "Уровень",
            "Глобальный блок",
            "Направление",
            "Система",
            "Групп",
            "Типов продукции",
            "Уникальных производителей",
            "Групп без производителей",
        ],
        summary_rows,
    )
    ws_summary.conditional_formatting.add(
        f"H2:H{ws_summary.max_row}",
        CellIsRule(operator="greaterThan", formula=["0"], fill=WARN_FILL),
    )

    review_rows = []
    for vendor in Vendor.objects.filter(status="manual_review_needed").order_by("name"):
        review_rows.append(
            [
                "Производитель",
                vendor.id,
                vendor.name,
                "",
                vendor.status,
                vendor.notes or vendor.source or "",
                "Проверить",
                "",
            ]
        )
    for link in VendorProductGroup.objects.filter(status="manual_review").select_related(
        "vendor", "product_group"
    ).order_by("vendor__name"):
        review_rows.append(
            [
                "Связь",
                link.id,
                link.vendor.name,
                link.product_group.title,
                link.status,
                link.notes or link.source or link.matched_by or "",
                "Проверить",
                "",
            ]
        )
    write_table(
        ws_review,
        [
            "Тип",
            "ID",
            "Производитель",
            "Товарная группа",
            "Статус",
            "Причина/источник",
            "Решение коммерческого отдела",
            "Комментарий",
        ],
        review_rows,
        editable_cols={7, 8},
        dropdowns={"Решение коммерческого отдела": action_values},
    )

    wb.save(out_file)

    check = load_workbook(out_file, read_only=True, data_only=True)
    print(out_file.resolve())
    for sheet in check.worksheets:
        print(f"{sheet.title}: {sheet.max_row} rows, {sheet.max_column} cols")
    print(f"empty_groups: {len(empty_rows)}")
    print(f"vendors: {Vendor.objects.count()}")
    print(f"links: {VendorProductGroup.objects.count()}")


if __name__ == "__main__":
    main()
